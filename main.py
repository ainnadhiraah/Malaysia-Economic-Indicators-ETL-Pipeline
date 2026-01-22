import yaml
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os

from scripts.dosm_scrapers import (
    PPI_Scraper,
    WholesaleRetail_Scraper,
    GDP_Scraper,
    Productivity_Scraper,
)

# Map name in YAML to scraper class
SCRAPER_MAP = {
    "PPI": PPI_Scraper,
    "WholesaleRetail": WholesaleRetail_Scraper,
    "GDP": GDP_Scraper,
    "Productivity": Productivity_Scraper,
}

def main():
    with open("config/sources.yaml", "r") as f:
        config = yaml.safe_load(f)

    merged_df = pd.DataFrame()

    for source in config["sources"]:
        name = source["name"]
        url = source["url"]
        last_updated = source["last_updated"]

        scraper_class = SCRAPER_MAP.get(name)
        if not scraper_class:
            print(f"No scraper defined for {name}")
            continue

        print(f"Fetching {name}...")
        scraper = scraper_class(url, last_updated)
        df = scraper.run()

        if df.empty:
            print(f"No new data for {name}")
            continue

        if merged_df.empty:
            merged_df = df
        else:
            merged_df = pd.merge(merged_df, df, on="Date", how="outer")

    # Sort by actual datetime and reformat
    merged_df["Date"] = pd.to_datetime(merged_df["Date"], errors="coerce")
    merged_df = merged_df.sort_values(by="Date")
    merged_df["Date"] = merged_df["Date"].dt.strftime("%b-%y")
    merged_df = merged_df.fillna("#N/A")

    os.makedirs("data", exist_ok=True)
    output_path = "data/Haver_data_output.xlsx"

    # --- CASE 1: File exists (refresh only data starting from row 4) ---
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb.active

        start_row = ws.max_row + 1

        # Append new data starting from that row
        for r_idx, row in enumerate(merged_df.itertuples(index=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_path)
        wb.close()
        print(f"Existing workbook updated")

    # --- CASE 2: File doesn't exist (create full new workbook) ---
    else:
        merged_df.to_excel(output_path, index=False, startrow=2)

        wb = load_workbook(output_path)
        ws = wb.active

        # Add source links in row 1
        sources = {
            "B1": "https://open.dosm.gov.my/data-catalogue/ppi",
            "C1": "https://open.dosm.gov.my/data-catalogue/iowrt",
            "D1": "https://open.dosm.gov.my/data-catalogue/gdp_qtr_real_demand?series=abs&type=e5&visual=table",
            "E1": "https://open.dosm.gov.my/data-catalogue/gdp_qtr_real_demand?series=abs&type=e6&visual=table",
            "F1": None,
            "G1": "https://open.dosm.gov.my/data-catalogue/productivity_qtr?series=abs&sector=p1&visual=table",
            "H1": "https://open.dosm.gov.my/data-catalogue/productivity_qtr?series=abs&sector=p2&visual=table",
            "I1": "https://open.dosm.gov.my/data-catalogue/productivity_qtr?series=abs&sector=p3&visual=table",
            "J1": "https://open.dosm.gov.my/data-catalogue/productivity_qtr?series=abs&sector=p4&visual=table",
            "K1": "https://open.dosm.gov.my/data-catalogue/productivity_qtr?series=abs&sector=p5&visual=table",
        }

        for cell, url in sources.items():
            if url:
                ws[cell].value = "OpenDOSM_Source"
                ws[cell].hyperlink = url
                ws[cell].style = "Hyperlink"
            else:
                ws[cell].value = "Derived (D - E)"

        # Add frequency info in row 2
        ws["B2"] = "Monthly"
        ws["C2"] = "Monthly"
        for col in ["D", "E", "F", "G", "H", "I", "J", "K"]:
            ws[f"{col}2"] = "Quarterly"

        wb.save(output_path)
        wb.close()
        print(f"New workbook created at {output_path}")

    print(f"Data successfully saved to {output_path}")


if __name__ == "__main__":
    main()
